import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

path = sys.argv[1]
# data start from A9
# columns: date, details, receipt number, amount (try), balance (try)
book = openpyxl.load_workbook(path)
sheet = book.active
for row in range(9, sheet.max_row + 1):
    operation = {"Account": "maha-deniz",
                 "DateTime": sheet.cell(row, 1).value,
                 "Description": sheet.cell(row, 2).value,
                 "Amount": str(sheet.cell(row, 4).value) + " try"}
    print(str(operation).replace("'", "\""))
