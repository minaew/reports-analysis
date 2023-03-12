import sys
import openpyxl
import fix_fill

sys.stdout.reconfigure(encoding='utf-8')
path = sys.argv[1]
with fix_fill.fix_fill(path) as new_path:
    # data start from A13
    # columns: date, invoice no, explanation, transaction amount, balance
    book = openpyxl.load_workbook(new_path)
    sheet = book.active
    for row in range(13, sheet.max_row - 7):
        operation = {"Account": "leha-ziraat",
                     "DateTime": sheet.cell(row, 1).value,
                     "Description": sheet.cell(row, 2).value,
                     "Amount": str(sheet.cell(row, 4).value) + " try"}
        print(str(operation).replace("'", "\""))
